from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
import cv2
import numpy as np
from io import BytesIO
import easyocr
from ultralytics import YOLO
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import tempfile
import shutil
import hashlib
from datetime import datetime

app = FastAPI()

# Enable CORS for frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # You can restrict this to your frontend domain in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load the YOLO model (ensure the path to your model is correct)
model = YOLO('backend/models/best.pt')

# Initialize the OCR reader globally
reader = easyocr.Reader(['en'])

# Create or load an Excel sheet to store results
def update_excel(plate_data):
    excel_file = 'backend/data/number_plate_results.xlsx'
    file_exists = Path(excel_file).exists()

    if file_exists:
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detected Plates"
        ws.cell(row=1, column=1, value="S.No")
        ws.cell(row=1, column=2, value="Number Plate Text")
        ws.cell(row=1, column=3, value="Date")
        ws.cell(row=1, column=4, value="Time")
        ws.cell(row=1, column=5, value="Count")

    for plate_text, data in plate_data.items():
        current_row = ws.max_row + 1
        ws.cell(row=current_row, column=1, value=current_row - 1)
        ws.cell(row=current_row, column=2, value=plate_text)
        ws.cell(row=current_row, column=3, value=data['date'])
        ws.cell(row=current_row, column=4, value=data['time'])
        ws.cell(row=current_row, column=5, value=data['count'])

    wb.save(excel_file)

def get_ocr(im, coors):
    x, y, w, h = int(coors[0]), int(coors[1]), int(coors[2]), int(coors[3])
    img_height, img_width, _ = im.shape
    x = max(0, x)
    y = max(0, y)
    w = min(img_width, w)
    h = min(img_height, h)

    roi = im[y:h, x:w]
    if roi.size == 0 or w <= x or h <= y:
        return ""

    gray = cv2.cvtColor(roi, cv2.COLOR_RGB2GRAY)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    results = reader.readtext(thresh)
    for result in results:
        text = result[1]
        confidence = result[2]
        if len(text) > 3 and confidence > 0.2:
            return text
    return ""

def process_frame(img):
    results = model(img)
    plate_data = {}
    for result in results:
        for box in result.boxes:
            coors = [int(box.xyxy[0][0]), int(box.xyxy[0][1]), int(box.xyxy[0][2]), int(box.xyxy[0][3])]
            ocr_text = get_ocr(img, coors)
            if ocr_text:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                date_str, time_str = timestamp.split(' ')
                if ocr_text not in plate_data:
                    plate_data[ocr_text] = {'date': date_str, 'time': time_str, 'count': 1}
                else:
                    plate_data[ocr_text]['count'] += 1
            cv2.rectangle(img, (coors[0], coors[1]), (coors[2], coors[3]), (0, 255, 0), 2)
            cv2.putText(img, ocr_text, (coors[0], coors[1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36, 255, 12), 2)
    return img, plate_data

@app.post("/detect-image/")
async def detect_image(file: UploadFile = File(...)):
    try:
        image_data = await file.read()
        nparr = np.frombuffer(image_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if img is None:
            raise HTTPException(status_code=400, detail="Invalid image file")

        processed_img, plate_data = process_frame(img)

        if plate_data:
            update_excel(plate_data)

        _, img_encoded = cv2.imencode('.jpg', processed_img)
        img_bytes = BytesIO(img_encoded.tobytes())
        return StreamingResponse(img_bytes, media_type="image/jpeg")

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/detect-video/")
async def detect_video(file: UploadFile = File(...)):
    try:
        video_data = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".mp4") as temp_video:
            temp_video.write(video_data)
            temp_video_path = temp_video.name

        video = cv2.VideoCapture(temp_video_path)
        if not video.isOpened():
            raise HTTPException(status_code=400, detail="Invalid video file")

        frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
        frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter('output.mp4', fourcc, 20.0, (frame_width, frame_height))

        all_plate_data = {}

        while True:
            ret, frame = video.read()
            if not ret:
                break

            processed_frame, plate_data = process_frame(frame)

            for plate, data in plate_data.items():
                if plate not in all_plate_data:
                    all_plate_data[plate] = {'date': data['date'], 'time': data['time'], 'count': data['count']}
                else:
                    all_plate_data[plate]['count'] += data['count']

            out.write(processed_frame)

        if all_plate_data:
            update_excel(all_plate_data)

        out.release()
        video.release()

        shutil.rmtree(temp_video_path, ignore_errors=True)

        return StreamingResponse(open('output.mp4', 'rb'), media_type="video/mp4")

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=5000)
